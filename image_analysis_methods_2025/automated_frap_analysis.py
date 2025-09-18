import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
import statistics
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
from PIL import Image
import tempfile

# define helper functions:
# function to trucate values from the excel file
def truncate(x, decimals):
    if isinstance(x, float):
        factor = 10 ** decimals
        return math.trunc(x * factor) / factor
    return x

# function to average values stored in a range of rows
def avg(ws, col, start_row, end_row):
    values = [ws[f"{col}{row}"].value for row in range(start_row, end_row + 1)]
    values = [v for v in values if isinstance(v, (int, float))]
    return statistics.mean(values) if values else None

# function to save graphs as jpg files - not included in the protocol
def save_all_to_jpg(figs, image_path, output_jpg):
    # 1) Render each Matplotlib Figure to a temporary PNG
    temp_paths = []
    for fig in figs:
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        fig.savefig(tmp.name, dpi=300, bbox_inches="tight")
        temp_paths.append(tmp.name)
        plt.close(fig)

    # 2) Add the inclusion image last
    temp_paths.append(image_path)

    # 3) Open all images and measure total size
    images = [Image.open(p) for p in temp_paths]
    widths, heights = zip(*(im.size for im in images))
    
    max_width = max(widths)
    total_height = sum(heights)
    
    # 4) Create a blank canvas and paste each image centered
    canvas = Image.new("RGB", (max_width, total_height), (255,255,255))
    y_offset = 0
    for im in images:
        x_offset = (max_width - im.width) // 2
        canvas.paste(im, (x_offset, y_offset))
        y_offset += im.height
    
    # 5) Save the combined canvas as a single JPEG
    canvas.save(output_jpg, "JPEG", quality=95)
    return output_jpg

# this is the same as what is included in the protocol, though the protcol doesn't outline it as a functinon
def run_frap_analysis(df: pd.DataFrame, time_increment: float, size_of_inclusion: float, image_name: str, bleaching_time: int, image_path: str):
    analysis = df.iloc[:, :3].copy()

    analysis_path   = f"{image_name}_analysis.xlsx"
    fit_curve_path  = f"{image_name}_fit_curve.xlsx"

    analysis.iloc[:, 0] = analysis.iloc[:, 0].apply(lambda x: truncate(x, 2))
    analysis.iloc[:, 1] = analysis.iloc[:, 1].apply(lambda x: truncate(x, 2))
    analysis.iloc[:, 2] = analysis.iloc[:, 2].apply(lambda x: truncate(x, 2))

    analysis = analysis.iloc[1:]

    # Step 3 (safe version): Add or replace cycle column
    analysis['cycle'] = list(range(1, len(analysis) + 1))

    # Reorder so 'cycle' is the first column
    analysis = analysis[['cycle'] + [col for col in analysis.columns if col != 'cycle']]

    cols = analysis.columns.tolist()

    cols[1] = 'ROI1'
    cols[2] = 'ROI2'
    cols[3] = 'ROI3'

    analysis.columns = cols

    num_rows = len(analysis)
    time_column = [round((i + 1) * time_increment, 2) for i in range(num_rows)]

    analysis["TIME (sec)"] = time_column

    analysis.insert(5, 'ROI1 ', [None] * num_rows)
    analysis.insert(6, 'ROI2 ', [None] * num_rows)
    analysis.insert(7, 'ROI3 ', [None] * num_rows)

    analysis.to_excel(f"{image_name}_analysis.xlsx", index=False)


    # Load the existing workbook and worksheet
    wb = load_workbook(f"{image_name}_analysis.xlsx")
    ws = wb.active  # or wb['SheetName'] if named

    # === BEGIN FORMATTING SECTION ===
    red_accent_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    orange_accent_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    aqua_accent_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    dark_blue_accent_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    purple_accent_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
    light_pink_accent_fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
    light_brown_accent_fill = PatternFill(start_color="C4BD97", end_color="C4BD97", fill_type="solid")
    light_green_accent_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    blue_accent_fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
    # === END FORMATTING SECTION

    f6 = avg(ws, "B", 2, bleaching_time)
    g6 = avg(ws, "C", 2, bleaching_time)
    h6 = avg(ws, "D", 2, bleaching_time)
    g7 = avg(ws, "C", bleaching_time + 1, 100) 

    
    j7 = g7 / g6 if g6 else None

    # Assign static labels
    ws["I6"] = f"AVERAGE {bleaching_time - 1} images before bleaching"
    ws["I7"] = "AVERAGE all images after bleaching"
    ws["J6"] = "r RO2"
    ws["K1"] = "normalized F ROI1"
    ws["L6"] = "plateau best fit value (prism)"
    ws["M6"] = "mobile fraction Fm"
    ws["N6"] = "immobile fraction Fi"
    ws["O6"] = "size of inclusion"
    ws["P6"] = "average normalized F ROI1 before bleaching"

    # Assign computed values
    ws["F6"] = f6
    ws["G6"] = g6
    ws["H6"] = h6
    ws["G7"] = g7
    ws["J7"] = j7
    ws["O7"] = size_of_inclusion

    # Compute normalized F ROI1 manually in K2:K{end}
    for row in range(2, ws.max_row + 1):
        b = ws[f"B{row}"].value
        d = ws[f"D{row}"].value
        if isinstance(b, (int, float)) and isinstance(d, (int, float)) and j7:
            ws[f"K{row}"] = (b - d) / j7
        else:
            ws[f"K{row}"] = None

    p7 = avg(ws, "K", 2, bleaching_time)
    ws["P7"] = p7

    # === BEGIN FORMATTING SECTION ===
    for cell in ["F6", "G6", "H6", "I6"]:
        ws[cell].fill = red_accent_fill

    for cell in ["G7", "H7", "I7"]:
        ws[cell].fill = orange_accent_fill  

    for cell in ["J6", "J7"]:
        ws[cell].fill = aqua_accent_fill

    for cell in ["L6", "L7"]:
        ws[cell].fill = purple_accent_fill

    for cell in ["M6", "M7"]:
        ws[cell].fill = light_pink_accent_fill

    for cell in ["N6", "N7"]:
        ws[cell].fill = light_brown_accent_fill

    for cell in ["O6", "O7"]:
        ws[cell].fill = light_green_accent_fill

    for cell in ["P6", "P7"]:
        ws[cell].fill = blue_accent_fill

    for row in range(1, ws.max_row + 1):
        ws[f"K{row}"].fill = dark_blue_accent_fill

    ws.row_dimensions[6].height = 57
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["L"].width = 12.57
    ws.column_dimensions["P"].width = 11.14


    ws["I6"].alignment = Alignment(wrap_text=True)
    ws["I7"].alignment = Alignment(wrap_text=True)
    ws["L6"].alignment = Alignment(wrap_text=True)
    ws["M6"].alignment = Alignment(wrap_text=True)
    ws["N6"].alignment = Alignment(wrap_text=True)
    ws["O6"].alignment = Alignment(wrap_text=True)
    ws["P6"].alignment = Alignment(wrap_text=True)

    column_indices = [1, 2, 3, 4, 5, 11]

    for col_idx in column_indices:
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

    wb.save(f"{image_name}_analysis.xlsx")
    # === END FORMATTING SECTION ===

    df = pd.read_excel(f"{image_name}_analysis.xlsx")
    x = df.iloc[bleaching_time-1:, 4].values   # Column E is index 4
    y = df.iloc[bleaching_time-1:, 10].values  # Column K is index 10

    def exp_plateau_fixed_y0(x, YM, k):
        return YM * (1 - np.exp(-k * x))

    # Fit the model
    popt, pcov = curve_fit(exp_plateau_fixed_y0, x, y, p0=[max(y), 0.01], bounds=([0, 1e-8], [np.inf, np.inf]))
    YM, k = popt

    # Generate fit line
    x_fit = np.linspace(min(x), max(x), 200)
    y_fit = exp_plateau_fixed_y0(x_fit, *popt) # will not include this unless things change

    plateu = YM
    plateu = round(plateu, 2)

    wb = load_workbook(f"{image_name}_analysis.xlsx")
    ws = wb.active 

    # Write the plateau value to the worksheet
    ws["L7"] = plateu
    m7 = plateu / p7 
    n7 = 1 - m7 
    ws["M7"] = m7
    ws["N7"] = n7

    wb.save(f"{image_name}_analysis.xlsx")

    column_e = df.iloc[:, 4].dropna()

    time_frame = column_e.iloc[-1]

    x_generated = np.linspace(0, time_frame, 1000)
    y_generated = exp_plateau_fixed_y0(x_generated, YM, k)

    # Create DataFrame and write to Excel
    fit_df = pd.DataFrame({
        "Time (sec)": x_generated,
        "Fluorescence intensity": y_generated, 
        "Fluorescence intensity (percentage)": (y_generated * 100) / p7
    })


    fit_df.to_excel(f"{image_name}_fit_curve.xlsx", index=False)


    # plot each of the graphs

    df = pd.read_excel(f"{image_name}_analysis.xlsx")

    df["cycle"] = pd.to_numeric(df["cycle"], errors="coerce")
    df["ROI1"]  = pd.to_numeric(df["ROI1"],  errors="coerce")
    df["ROI2"]  = pd.to_numeric(df["ROI2"],  errors="coerce")
    df["ROI3"]  = pd.to_numeric(df["ROI3"],  errors="coerce")
    # Get all rows from the first 4 columns
    df = df.iloc[:, :4]
    df.columns = ["cycle", "ROI1", "ROI2", "ROI3"]

    # 1) RAW DATA
    fig1 = plt.figure(figsize=(6,4))
    ax1  = fig1.subplots()
    x_vals = df["cycle"].to_numpy(dtype=float)
    y1     = df["ROI1"].to_numpy(dtype=float)
    y2     = df["ROI2"].to_numpy(dtype=float)
    y3     = df["ROI3"].to_numpy(dtype=float)
    ax1.plot(x_vals, y1, 'o-', label="ROI1", color="red", linewidth=1, markersize=5)
    ax1.plot(x_vals, y2, 'o-', label="ROI2", color="blue", linewidth=1, markersize=5)
    ax1.plot(x_vals, y3, 'o-', label="ROI3", color="green", linewidth=1, markersize=5)
    ax1.set(xlabel="cycle", ylabel="Fluorescence Intensity (a.u.)",
            title=f"{image_name} RAW DATA",
            xlim=(0, x_vals.max()+2),
            ylim=(-0.7, max(y1.max(),y2.max(),y3.max())*1.15))
    ax1.legend(loc="center left", bbox_to_anchor=(1,0.5), fontsize=9)
    fig1.tight_layout()

    # 2) NORMALIZED DATA
    fig2 = plt.figure(figsize=(6,4))
    ax2  = fig2.subplots()
    ax2.plot(x, y, '-', color='black', linewidth=1)
    ax2.plot(x, y, 'o', color='red', markersize=5)
    ax2.plot(x_generated, y_generated, '-', color='red', label='Fit curve')
    ax2.set(xlabel="Time (sec)", ylabel="Normalized Fluorescence Intensity (a.u.)",
            title=f"{image_name} NORMALIZED DATA")
    ax2.set(xlim=(0,None), ylim=(0,None))
    ax2.legend()
    fig2.tight_layout()

    # 3) FIT CURVE PERCENTAGE
    fig3 = plt.figure(figsize=(6,4))
    ax3  = fig3.subplots()
    ax3.plot(x_generated, fit_df["Fluorescence intensity (percentage)"],
             '-', color='black', linewidth=3, label='Fit curve')
    ax3.axhline(100, linestyle='--', color ='black', linewidth=1, label='100%')
    ax3.set(xlabel="Time (sec)", ylabel="Normalized Fluorescence Intensity (%)",
            title=f"{image_name} FIT CURVE PERCENTAGE")
    ax3.set(xlim=(0,None), ylim=(0,None))
    ax3.legend()
    fig3.tight_layout()


    jpg_file = save_all_to_jpg([fig1, fig2, fig3], image_path, f"{image_name}_complete.jpg")
    

    return analysis_path, fit_curve_path, [fig1, fig2, fig3], jpg_file